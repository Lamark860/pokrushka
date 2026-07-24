"""Чтение выгрузок статистики: CSV и XLSX.

Точный формат выгрузки Дзена нам пока неизвестен (ждём файл от Артура), поэтому парсер
намеренно терпимый: сам определяет кодировку и разделитель, понимает русские числа
(«1 234», «12,5%») и угадывает колонки по синонимам. Что не угадалось — пользователь
сопоставляет руками в интерфейсе.
"""
import csv
import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime

# Поля, которые нас интересуют, и синонимы заголовков (нижний регистр, без пунктуации)
FIELD_SYNONYMS: dict[str, tuple[str, ...]] = {
    "title": ("заголовок", "название", "публикация", "статья", "материал", "title"),
    "url": ("ссылка", "адрес", "url", "link", "ссылка на публикацию"),
    "views": ("показы", "просмотры", "показов", "просмотров", "views", "impressions"),
    "reads": ("дочитывания", "дочитываний", "прочтения", "дочитывания шт", "reads"),
    "read_ratio": ("дочитываемость", "процент дочитываний", "дочитывания %", "read ratio"),
    "likes": ("лайки", "нравится", "likes"),
    "comments": ("комментарии", "комментариев", "comments"),
    "reposts": ("репосты", "поделились", "shares", "reposts"),
    "collected_at": ("дата", "день", "период", "date"),
}

NUMERIC_FIELDS = ("views", "reads", "likes", "comments", "reposts")

_PUNCT = re.compile(r"[^\w\s%]+", re.UNICODE)
_SPACES = re.compile(r"\s+")
# Неразрывный и узкий неразрывный пробелы — обычные гости в выгрузках
_NBSP = "   "


class TableError(ValueError):
    """Файл не читается или в нём нет заголовков."""


@dataclass
class Table:
    headers: list[str]
    rows: list[dict[str, str]]
    mapping: dict[str, str] = field(default_factory=dict)  # поле → заголовок колонки

    @property
    def unmapped_fields(self) -> list[str]:
        return [name for name in FIELD_SYNONYMS if name not in self.mapping]


def normalize_header(value: str) -> str:
    cleaned = _PUNCT.sub(" ", (value or "").strip().lower())
    return _SPACES.sub(" ", cleaned).strip()


def guess_mapping(headers: list[str]) -> dict[str, str]:
    """Сопоставить наши поля с колонками файла по синонимам заголовков."""
    normalized = {normalize_header(h): h for h in headers if h and h.strip()}
    mapping: dict[str, str] = {}
    taken: set[str] = set()

    for field_name, synonyms in FIELD_SYNONYMS.items():
        # сначала точное совпадение, потом вхождение синонима в заголовок
        for synonym in synonyms:
            header = normalized.get(synonym)
            if header and header not in taken:
                mapping[field_name] = header
                taken.add(header)
                break
        else:
            for norm, header in normalized.items():
                if header in taken:
                    continue
                if any(synonym in norm for synonym in synonyms):
                    mapping[field_name] = header
                    taken.add(header)
                    break
    return mapping


def parse_number(value: str | int | float | None) -> int | None:
    """«1 234» → 1234, «12,5» → 12 (округление), пусто и мусор → None."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(round(value))

    text = str(value).strip()
    for space in _NBSP:
        text = text.replace(space, "")
    text = text.replace(" ", "").replace("%", "").replace(",", ".")
    if not text or not re.fullmatch(r"-?\d+(\.\d+)?", text):
        return None
    return int(round(float(text)))


def parse_ratio(value: str | int | float | None) -> float | None:
    """Дочитываемость: «42,5%» → 42.5, «0,425» → 42.5."""
    if value is None or isinstance(value, bool):
        return None
    text = str(value).strip()
    for space in _NBSP:
        text = text.replace(space, "")
    has_percent = "%" in text
    text = text.replace(" ", "").replace("%", "").replace(",", ".")
    if not text or not re.fullmatch(r"-?\d+(\.\d+)?", text):
        return None
    number = float(text)
    # Доля вида 0.42 — приводим к процентам, если знака процента в ячейке не было
    if not has_percent and 0 <= number <= 1:
        number *= 100
    return number


_DATE_FORMATS = ("%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y", "%Y/%m/%d", "%d-%m-%Y")


def parse_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    # «01.07.2026 - 07.07.2026» — берём конец периода: срез описывает состояние на эту дату
    if "-" in text and text.count(".") >= 2:
        text = text.split("-")[-1].strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


# ------------------------------------------------------------------ чтение файлов


def _decode(payload: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            return payload.decode(encoding)
        except UnicodeDecodeError:
            continue
    return payload.decode("utf-8", errors="replace")


def _sniff_delimiter(sample: str) -> str:
    try:
        return csv.Sniffer().sniff(sample, delimiters=";\t,").delimiter
    except csv.Error:
        # Выгрузки Дзена/vc — обычно «;», это и берём как запасной вариант
        counts = {sep: sample.count(sep) for sep in ";\t,"}
        return max(counts, key=counts.get) if any(counts.values()) else ";"


def read_csv(payload: bytes) -> Table:
    text = _decode(payload)
    if not text.strip():
        raise TableError("Файл пустой")

    delimiter = _sniff_delimiter(text[:4096])
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = [row for row in reader if any(cell.strip() for cell in row)]
    if not rows:
        raise TableError("В файле нет строк")

    headers = [cell.strip() for cell in rows[0]]
    body = [
        {headers[i]: (row[i].strip() if i < len(row) else "") for i in range(len(headers))}
        for row in rows[1:]
    ]
    return Table(headers=headers, rows=body, mapping=guess_mapping(headers))


def read_xlsx(payload: bytes) -> Table:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - зависимость объявлена в requirements
        raise TableError("Не установлен openpyxl — XLSX читать нечем") from exc

    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)

    headers: list[str] = []
    for row in rows_iter:
        if row and any(cell is not None and str(cell).strip() for cell in row):
            headers = [str(cell).strip() if cell is not None else "" for cell in row]
            break
    if not headers:
        raise TableError("В файле нет заголовков")

    body: list[dict[str, str]] = []
    for row in rows_iter:
        if not row or not any(cell is not None and str(cell).strip() for cell in row):
            continue
        body.append(
            {
                headers[i]: ("" if i >= len(row) or row[i] is None else str(row[i]).strip())
                for i in range(len(headers))
            }
        )
    workbook.close()
    return Table(headers=headers, rows=body, mapping=guess_mapping(headers))


def read_table(filename: str, payload: bytes) -> Table:
    if filename.lower().endswith((".xlsx", ".xlsm")):
        return read_xlsx(payload)
    return read_csv(payload)
